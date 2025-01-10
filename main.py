import win32print
import time
from openpyxl import Workbook, load_workbook
from datetime import datetime

register_path = r"\\192.168.2.8\isos\Relatorios\arquivo_com_data_e_hora.xlsx"
printer_name = "4BARCODE 4B-2082A (Copiar 4)"
rather_document = "Fast Report Document"

# Função para inicializar ou carregar o arquivo Excel
def init_excel():
    try:
        # Tentar carregar o arquivo existente
        workbook = load_workbook(register_path)
        sheet = workbook.active
    except FileNotFoundError:
        # Se o arquivo não existe, criar um novo
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Registros de Impressão"
        sheet.append(["Data/Hora", "ID do Trabalho", "Documento", "Usuário", "Status"])
        workbook.save(register_path)
    return workbook, sheet

# Função para carregar os IDs já registrados
def load_id(sheet):
    ids_registrados = set()
    for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
        ids_registrados.add(row[0])  # Adicionar o ID do trabalho na lista
    return ids_registrados

# Abrir conexão com a impressora

def printer_monitor (printer_name_in):
    
    try:
        nPrinter = win32print.OpenPrinter(printer_name_in)
        workbook, sheet = init_excel()
        ids_registrados = load_id(sheet)

        # Obter informações sobre a impressora
        #printer_info = win32print.GetPrinter(hPrinter, 2)

        while True:
            print("-----------------------Aguardando trabalho de impressão--------------------------")
            jobs = win32print.EnumJobs(nPrinter, 0, -1, 2)  
            if jobs:
                for job in jobs: 
                    job_id = job['JobId']
                    documento = job['pDocument']
                    
                    if documento == rather_document:
                        if job_id not in ids_registrados: #verifica ids
                            #Salva no documento
                            data_hora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            documento = job['pDocument']
                            usuario = job['pUserName']
                            status = job.get('Status', 'Desconhecido')

                            #Exibe na tela
                            print(f"\nPáginas Impressas: {job['PagesPrinted']}")
                            print(f"ID do Trabalho: {job['JobId']}")
                            print(f"Status: {job['Status']}")
                            print(f"Documento: {job['pDocument']}\n")

                            # Salvar no Excel
                            sheet.append([data_hora, job_id, documento, usuario, status])
                            ids_registrados.add(job_id)
                            workbook.save(register_path)    
            else:
                print(f"Sem trabalhos na fila para a impressora '{printer_name_in}'.")
            time.sleep(5)
           
        
    except Exception as e:
        print(f"erro")

# Fechar conexão
    finally:
        win32print.ClosePrinter(nPrinter)
        #workbook.save(register_path)

if __name__ == '__main__':
    printer_monitor(printer_name)